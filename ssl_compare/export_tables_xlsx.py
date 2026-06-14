"""Export the 3-fold grouped-CV result tables to a single .xlsx workbook.

Sheets:
  Run67_accuracy / Run67_f1   -> all 9 tags (Run6 fold0 + Run7 fold1/2)
  DAPT_acc / DAPT_f1          -> dapt ablation (baseline / mask0.35 / +aug) + R-GRU

Values are 3-fold mean and std, per label_rate. Reads only existing result CSVs
(reuses the same loaders as the plotting scripts); fabricates nothing.
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment

import plot_run67_3fold as p67
import plot_dapt_ablation as pda

OUT_XLSX = os.path.join(p67.OUT_DIR, "ssl_results_tables.xlsx")

HEADER = ["tag", "0.01 mean", "0.01 std", "0.02 mean", "0.02 std",
          "0.05 mean", "0.05 std"]


def _write_sheet(ws, header, title, row_iter):
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append(header)
    for c in ws[2]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    for label, means, stds in row_iter:
        row = [label]
        for i in range(len(means)):
            row += [round(float(means[i]), 4), round(float(stds[i]), 4)]
        ws.append(row)
    ws.column_dimensions["A"].width = 26
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 11


def run67_rows(metric):
    rows = p67.load_raw_rows()
    for tag in p67.ALL_TAGS:
        _, m, s = p67.aggregate(rows, tag, metric)
        yield (tag, m, s)


def dapt_rows(metric):
    for v in pda.VARIANTS:
        rows = pda.load_raw_rows(v["raw_csvs"])
        for sub, lbl in [("dapt_ft", "%s [ft]"), ("dapt", "%s [frozen]")]:
            _, m, s = pda.aggregate(rows, sub, metric)
            yield (lbl % v["label"], m, s)
    rg = pda.load_raw_rows(pda.RGRU_CSVS)
    _, m, s = pda.aggregate(rg, "R-GRU", metric)
    yield ("R-GRU (baseline)", m, s)


def main():
    os.makedirs(p67.OUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    specs = [
        ("Run67_accuracy", "Run6/7 grouped 3-fold CV - Accuracy (mean / std)", run67_rows("acc")),
        ("Run67_f1",       "Run6/7 grouped 3-fold CV - Macro-F1 (mean / std)", run67_rows("f1")),
        ("DAPT_acc",       "DAPT ablation - Accuracy (mean / std)", dapt_rows("acc")),
        ("DAPT_f1",        "DAPT ablation - Macro-F1 (mean / std)", dapt_rows("f1")),
    ]
    for name, title, it in specs:
        ws = wb.create_sheet(name)
        _write_sheet(ws, HEADER, title, it)

    wb.save(OUT_XLSX)
    print("[saved] %s" % OUT_XLSX)


if __name__ == "__main__":
    main()
